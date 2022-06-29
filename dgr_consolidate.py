import datetime
import os
import shutil
import xlrd
from sqlalchemy import create_engine
import pandas as pd
import openpyxl
import warnings

warnings.filterwarnings("ignore")

db_connection = create_engine('mysql+mysqldb://root:rutu@localhost/TEIM')


def create_tables():
    db_connection.execute('CREATE TABLE IF NOT EXISTS generation_mastersheet(' +
                          'date DATE,' +
                          'financial_year VARCHAR(10),' +
                          'customer_name VARCHAR(255),' +
                          'client_name VARCHAR(255),' +
                          'state VARCHAR(255),' +
                          'site_name VARCHAR(255),' +
                          'wind_turbine_location_number VARCHAR(20),' +
                          'day_generation_kwh FLOAT,' +
                          'day_generation_hours FLOAT,' +
                          'operating_hours FLOAT,' +
                          'machine_availability_percent FLOAT,' +
                          'internal_grid_availability_percent FLOAT,' +
                          'external_grid_availability_percent FLOAT,' +
                          'plant_load_factor FLOAT,' +
                          'nor FLOAT,' +
                          'force_majeure FLOAT,' +
                          'grid_failure FLOAT,' +
                          'internal_grid_failure_hours FLOAT,' +
                          'unscheduled_services FLOAT,' +
                          'scheduled_services FLOAT)')

    db_connection.execute('CREATE TABLE IF NOT EXISTS breakdown_mastersheet(' +
                          'date DATE,' +
                          'financial_year VARCHAR(10),' +
                          'customer_name VARCHAR(255),' +
                          'client_name VARCHAR(255),' +
                          'state VARCHAR(255),' +
                          'site_name VARCHAR(255),' +
                          'wind_turbine_location_number VARCHAR(20),' +
                          'breakdown_remark VARCHAR(500),' +
                          'breakdown_hours FLOAT)')


def log_files(input_path):
    file_log = open('Logs/file_log.txt', 'a')

    files = [(name[0], filename) for name in os.walk(input_path) for filename in name[2]]

    for file in files:
        file_log.write(str(file) + '\t' + str(datetime.datetime.now()) + '\n')

    return files


def cleanup(input_path):
    destination = 'E:/SDA/DGR/'
    clients = ['INOX', 'REGEN', 'SENVION', 'SUZLON', 'TS WIND', 'WIND WORLD']

    for client in clients:
        files = [(name[0], filename) for name in os.walk(os.path.join(input_path, client)) for filename in name[2]]

        for file in files:
            shutil.move(file[0] + '/' + file[1], destination + client + '/' + file[1])


def timestamp_to_hours(sheet, from_column, to_column):
    for x in range(len(sheet)):
        if len(str(sheet.loc[x, from_column])) > 8:
            sheet.loc[x, to_column] = 24.0
        elif len(str(sheet.loc[x, from_column])) == 8:
            temp = list(map(float, str(sheet.loc[x, from_column]).split(':')))
            sheet.loc[x, to_column] = temp[0] + temp[1] / 60 + temp[2] / 3600
        else:
            sheet.loc[x, to_column] = 0.0

    return sheet


def timestamp_to_hours2(sheet, column):
    for x in range(len(sheet)):
        temp = list(map(float, str(sheet.loc[x, column]).split(':')))
        sheet.loc[x, column] = temp[0] + temp[1] / 60
    return sheet


def add_time(time1, time2):
    l1 = list(map(int, time1.split(':')))
    l2 = list(map(int, time2.split(':')))
    minutes = (l1[1] + l2[1]) % 60
    hours = l1[0] + l2[0] + int((l1[1] + l2[1]) / 60)
    return str(hours) + ':' + str(minutes)


def get_inox_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/INOX/'

    files = log_files(input_path)

    for file in files:

        if file[1].count('Generation') == 1:
            if file[1].endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
                generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0],
                                           skiprows=[0])
            else:
                workbook = xlrd.open_workbook(file[0] + '/' + file[1])
                generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0],
                                           skiprows=[0])

            generation = generation.drop(index=len(generation) - 1)

            generation['Date'] = pd.to_datetime(generation['Date'])
            generation['date'] = generation['Date'].dt.date
            generation['financial_year'] = [
                'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(
                    y - 2000)
                for x, y in
                zip(generation['Date'].dt.month, generation['Date'].dt.year)]
            generation['client_name'] = ['INOX'] * len(generation)

            generation = generation.reset_index().rename(columns={'Customer': 'customer_name',
                                                                  'States': 'state', 'Site': 'site_name',
                                                                  'Location No.': 'wind_turbine_location_number',
                                                                  'KWH': 'day_generation_kwh',
                                                                  'MA %': 'machine_availability_percent',
                                                                  'PLF %': 'plant_load_factor',
                                                                  'Production Hrs': 'day_generation_hours',
                                                                  'Operating Hrs': 'operating_hours'})

            generation = timestamp_to_hours2(generation, 'day_generation_hours')
            generation = timestamp_to_hours2(generation, 'operating_hours')

            generation['nor'] = [0.0] * len(generation)
            generation['force_majeure'] = [0.0] * len(generation)
            generation['internal_grid_failure_hours'] = [0.0] * len(generation)
            generation['grid_failure'] = [0.0] * len(generation)
            generation['unscheduled_services'] = [0.0] * len(generation)

            generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                     'state', 'site_name', 'wind_turbine_location_number',
                                     'day_generation_kwh', 'day_generation_hours',
                                     'operating_hours', 'machine_availability_percent',
                                     'plant_load_factor', 'nor', 'force_majeure', 'internal_grid_failure_hours',
                                     'grid_failure', 'unscheduled_services']]

            gen_data = pd.concat([gen_data, generation], ignore_index=True)

        else:
            if file[1].endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
                breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0],
                                          skiprows=[0])
            else:
                workbook = xlrd.open_workbook(file[0] + '/' + file[1])
                breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0],
                                          skiprows=[0])

            breakdown['Date'] = pd.to_datetime(breakdown['Date'])
            breakdown['date'] = breakdown['Date'].dt.date
            breakdown['financial_year'] = [
                'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(
                    y - 2000)
                for x, y in
                zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
            breakdown['client_name'] = ['INOX'] * len(breakdown)

            breakdown = timestamp_to_hours(breakdown, 'Total Stop', 'breakdown_hours')

            breakdown = breakdown.reset_index().rename(columns={'Customer': 'customer_name',
                                                                'State': 'state', 'Site': 'site_name',
                                                                'Location No.': 'wind_turbine_location_number',
                                                                'Reason': 'breakdown_remark'})

            bd_factors = {'BOC': 'nor', 'FORCE MAJEURE': 'force_majeure', 'GA INT': 'internal_grid_failure_hours',
                          'GA EXT': 'grid_failure', 'WTG': 'unscheduled_services'}

            temp_df = gen_data[gen_data['client_name'] == 'INOX']

            for x in range(len(breakdown)):
                if not breakdown.loc[x, 'Stop Due To'] == 'ACTIVITY BEYOND O&M CONTRACT':
                    df = temp_df[temp_df['date'] == breakdown.loc[x, 'date']]
                    df = df[
                        df['wind_turbine_location_number'] == breakdown.loc[x, 'wind_turbine_location_number']]
                    gen_data.loc[df.index.values, bd_factors[breakdown.loc[x, 'Stop Due To']]] += breakdown.loc[
                        x, 'breakdown_hours']

            breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                                   'state', 'site_name', 'wind_turbine_location_number',
                                   'breakdown_remark', 'breakdown_hours']]

            bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_regen_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/REGEN/'

    files = log_files(input_path)

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0, 1])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0], skiprows=[0, 1])

        loc_info = sheet.iloc[0][3].split(',')
        customer = loc_info[1].strip()
        site = loc_info[2].strip()
        state = loc_info[3].strip('.')
        date = sheet.iloc[1][3]

        sheet.columns = ['Project Location', 'wind_turbine_location_number', 'Cumulative Generation FinYr(KWH)',
                         'Cumulative Generation FinYr(HRS)', 'Monthly Cumulative Generation', 'Monthly PLF',
                         'day_generation_kwh',
                         'day_generation_hours', 'plant_load_factor', 'Lull Hrs', 'unscheduled_services',
                         'scheduled_services',
                         'machine_availability_percent', 'Internal Grid Fault Hrs', 'Internal Grid Shut down Hrs',
                         'External Grid Fault / S/d Hrs', 'External Grid Load Shedding',
                         'internal_grid_availability_percent',
                         'external_grid_availability_percent', 'force_majeure']

        generation = sheet.loc[3:17, :]
        generation = generation.reset_index().drop(columns=['index'])

        generation['Date'] = [date] * len(generation)
        generation['Date'] = pd.to_datetime(generation['Date'])
        generation['date'] = generation['Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(generation['Date'].dt.month, generation['Date'].dt.year)]
        generation['client_name'] = ['ReGen'] * len(generation)
        generation['customer_name'] = [customer] * len(generation)
        generation['state'] = [state] * len(generation)
        generation['site_name'] = [site] * len(generation)
        generation['internal_grid_failure_hours'] = generation['Internal Grid Fault Hrs'] + generation[
            'Internal Grid Shut down Hrs']
        generation['grid_failure'] = generation['External Grid Fault / S/d Hrs'] + generation[
            'External Grid Load Shedding']

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'state', 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'machine_availability_percent', 'internal_grid_availability_percent',
                                 'external_grid_availability_percent', 'plant_load_factor', 'grid_failure',
                                 'force_majeure', 'scheduled_services', 'unscheduled_services',
                                 'internal_grid_failure_hours']]

        breakdown = sheet.iloc[24:, [3, 4, 12]]
        breakdown.columns = ['wind_turbine_location_number', 'breakdown_remark', 'breakdown_hours']
        breakdown = breakdown.reset_index().drop(columns=['index'])

        breakdown['Date'] = [date] * len(breakdown)
        breakdown['Date'] = pd.to_datetime(breakdown['Date'])
        breakdown['date'] = breakdown['Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
        breakdown['client_name'] = ['ReGen'] * len(breakdown)
        breakdown['customer_name'] = [customer] * len(breakdown)
        breakdown['state'] = [state] * len(breakdown)
        breakdown['site_name'] = [site] * len(breakdown)

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'state', 'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_senvion_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/SENVION/'

    files = log_files(input_path)

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0])
            breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[1])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0],
                                       skiprows=[0])
            breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[1])

        generation['Date'] = pd.to_datetime(generation['Date'])
        generation['date'] = generation['Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in zip(generation['Date'].dt.month, generation['Date'].dt.year)]
        generation['client_name'] = ['SENVION'] * len(generation)
        file_name = file[1].split('-')
        generation['customer_name'] = [file_name[2].split('_')[0]] * len(generation)
        generation['site_name'] = [file_name[1]] * len(generation)

        generation = timestamp_to_hours(generation, 'Production hours', 'day_generation_hours')
        generation = timestamp_to_hours(generation, 'External Grid Down Time', 'grid_failure')
        generation = timestamp_to_hours(generation, 'Stoppage due to Customer / Utility Account/ Force major',
                                        'force_majeure')
        generation = timestamp_to_hours(generation, 'Scheduled services', 'scheduled_services')
        generation = timestamp_to_hours(generation, 'Unscheduled services', 'unscheduled_services')
        generation = timestamp_to_hours(generation, 'Enviromental', 'nor')
        generation = timestamp_to_hours(generation, 'Internal Grid Down Time', 'internal_grid_failure_hours')

        generation = generation.reset_index().rename(columns={'LOCATION  NO.': 'wind_turbine_location_number',
                                                              'WTG Production kWh ': 'day_generation_kwh',
                                                              'WTG Availability%': 'machine_availability_percent',
                                                              'Int. Grid Availability%': 'internal_grid_availability_percent',
                                                              'Ext. Grid Availability%': 'external_grid_availability_percent',
                                                              'PLF (WTG efficiency) %': 'plant_load_factor'})

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'machine_availability_percent', 'internal_grid_availability_percent',
                                 'external_grid_availability_percent', 'plant_load_factor',
                                 'grid_failure', 'force_majeure', 'nor', 'internal_grid_failure_hours',
                                 'scheduled_services', 'unscheduled_services']]

        breakdown['Date'] = pd.to_datetime(breakdown['Date'])
        breakdown['date'] = breakdown['Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(breakdown['Date'].dt.month, breakdown['Date'].dt.year)]
        breakdown['client_name'] = ['Senvion'] * len(breakdown)
        breakdown['customer_name'] = [file_name[2].split('_')[0]] * len(breakdown)
        breakdown['site_name'] = [file_name[1]] * len(breakdown)

        breakdown = timestamp_to_hours(breakdown, 'Total Duration', 'breakdown_hours')

        breakdown = breakdown.reset_index().rename(columns={'Loc No': 'wind_turbine_location_number',
                                                            'Error Description': 'breakdown_remark'})

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_suzlon_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/SUZLON/'

    files = log_files(input_path)

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0])
            breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[1])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            generation = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0])
            breakdown = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[1])

        generation['Gen. Date'] = pd.to_datetime(generation['Gen. Date'])
        generation['date'] = generation['Gen. Date'].dt.date
        generation['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(generation['Gen. Date'].dt.month, generation['Gen. Date'].dt.year)]
        generation['client_name'] = ['Suzlon'] * len(generation)

        generation.replace('*', 0.0, inplace=True)
        generation.replace('**', 0.0, inplace=True)

        generation = generation.reset_index().rename(columns={'Customer Name': 'customer_name',
                                                              'State': 'state', 'Site': 'site_name',
                                                              'Loc. No.': 'wind_turbine_location_number',
                                                              'Gen. (kwh) DAY': 'day_generation_kwh',
                                                              'Gen Hrs.': 'day_generation_hours',
                                                              'Opr Hrs.': 'operating_hours',
                                                              'M/C Avail.%': 'machine_availability_percent',
                                                              '%PLF DAY': 'plant_load_factor',
                                                              'GF': 'grid_failure',
                                                              'FM': 'force_majeure',
                                                              'S': 'scheduled_services',
                                                              'U': 'unscheduled_services'})

        generation = generation[['date', 'financial_year', 'customer_name', 'client_name',
                                 'state', 'site_name', 'wind_turbine_location_number',
                                 'day_generation_kwh', 'day_generation_hours',
                                 'operating_hours', 'machine_availability_percent',
                                 'plant_load_factor', 'grid_failure', 'force_majeure',
                                 'scheduled_services', 'unscheduled_services']]

        breakdown['Gen. Date'] = pd.to_datetime(breakdown['Gen. Date'])
        breakdown['date'] = breakdown['Gen. Date'].dt.date
        breakdown['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(breakdown['Gen. Date'].dt.month, breakdown['Gen. Date'].dt.year)]
        breakdown['client_name'] = ['Suzlon'] * len(breakdown)

        breakdown = breakdown.reset_index().rename(columns={'Customer Name': 'customer_name',
                                                            'State': 'state', 'Site': 'site_name',
                                                            'Loc. No.': 'wind_turbine_location_number',
                                                            'Breakdown Remark': 'breakdown_remark',
                                                            'Breakdown Hrs.': 'breakdown_hours'})

        breakdown = breakdown[['date', 'financial_year', 'customer_name', 'client_name',
                               'state', 'site_name', 'wind_turbine_location_number',
                               'breakdown_remark', 'breakdown_hours']]

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_tswind_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/TS WIND/'

    files = log_files(input_path)

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0, 1, 3])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0],
                                  skiprows=[0, 1, 3])

        sheet['DATE'] = pd.to_datetime(sheet['DATE'])
        sheet['date'] = sheet['DATE'].dt.date
        sheet['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(sheet['DATE'].dt.month, sheet['DATE'].dt.year)]
        sheet['client_name'] = ['TS Wind'] * len(sheet)

        sheet = timestamp_to_hours(sheet, 'PROD.', 'day_generation_hours')
        sheet = timestamp_to_hours(sheet, 'GRID DROP', 'grid_failure')
        sheet = timestamp_to_hours(sheet, 'FM.', 'force_majeure')
        sheet = timestamp_to_hours(sheet, 'MAINT.', 'scheduled_services')
        sheet = timestamp_to_hours(sheet, 'ERROR', 'unscheduled_services')
        sheet['breakdown_hours'] = sheet['grid_failure'] + sheet['force_majeure'] + sheet['scheduled_services'] + sheet[
            'unscheduled_services']

        sheet = sheet.reset_index().rename(columns={'CUSTOMER': 'customer_name', 'SITE': 'site_name',
                                                    'WTG': 'wind_turbine_location_number',
                                                    'GEN(FTD)': 'day_generation_kwh',
                                                    'MA(FTD)': 'machine_availability_percent',
                                                    'PLF': 'plant_load_factor', 'ERROR DETAILS': 'breakdown_remark'})

        generation = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                            'site_name', 'wind_turbine_location_number',
                            'day_generation_kwh', 'day_generation_hours',
                            'machine_availability_percent', 'plant_load_factor',
                            'grid_failure', 'force_majeure', 'scheduled_services', 'unscheduled_services']]

        breakdown = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                           'site_name', 'wind_turbine_location_number', 'breakdown_remark', 'breakdown_hours']]

        breakdown.dropna(inplace=True)

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def get_windworld_data(gen_data, bd_data):
    input_path = 'E:/SDA/Input DGR/WIND WORLD/'

    files = log_files(input_path)

    for file in files:
        if file[1].endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file[0] + '/' + file[1], read_only=True, keep_links=False)
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheetnames[0], skiprows=[0])
        else:
            workbook = xlrd.open_workbook(file[0] + '/' + file[1])
            sheet = pd.read_excel(os.path.join(file[0], file[1]), sheet_name=workbook.sheet_names()[0], skiprows=[0])

        sheet['DATE'] = pd.to_datetime(sheet['DATE'])
        sheet['date'] = sheet['DATE'].dt.date
        sheet['financial_year'] = [
            'FY ' + str(y) + '-' + str(y + 1 - 2000) if x not in [1, 2, 3] else 'FY ' + str(y - 1) + '-' + str(y - 2000)
            for x, y in
            zip(sheet['DATE'].dt.month, sheet['DATE'].dt.year)]
        sheet['client_name'] = ['Wind World'] * len(sheet)

        sheet = timestamp_to_hours(sheet, 'O.Hrs', 'operating_hours')

        bd_factors = {'BM': 'unscheduled_services', 'BD': 'unscheduled_services', 'GF': 'grid_failure',
                      'GS': 'grid_failure', 'PM': 'scheduled_services', 'SD': 'scheduled_services',
                      'FM': 'force_majeure', 'CS': 'grid_failure', 'LR': 'grid_failure', 'RF': 'force_majeure'}

        factors = ('BM', 'BD', 'GF', 'GS', 'PM', 'SD', 'FM', 'CS', 'LR', 'S/D', 'STS', 'RF')

        sheet['grid_failure'] = ['00:00'] * len(sheet)
        sheet['force_majeure'] = ['00:00'] * len(sheet)
        sheet['scheduled_services'] = ['00:00'] * len(sheet)
        sheet['unscheduled_services'] = ['00:00'] * len(sheet)
        sheet['breakdown_hours'] = ['00:00'] * len(sheet)

        sheet.fillna('', inplace=True)

        for x in range(len(sheet)):
            remark = sheet.loc[x, 'REMARKS']
            count = remark.lower().count('hrs')
            if count > 0:
                for z in range(count):
                    index = remark.lower().find('hrs')
                    time_str = remark[index - 6:index]
                    time_str = time_str.replace('.', ':')
                    i = time_str.find(':')
                    if i == -1:
                        time = time_str.strip()[-2:] + ':00'
                    elif time_str[i - 2:i].isnumeric():
                        time = time_str[i - 2:i + 3]
                    else:
                        time = time_str[i - 1:i + 3]
                    if z == 0 and not remark[:1].isdigit():
                        sheet.loc[x, 'breakdown_hours'] = add_time(sheet.loc[x, 'breakdown_hours'], time)
                        if remark[:3] == 'STS':
                            sheet.loc[x, 'force_majeure'] = add_time(sheet.loc[x, 'force_majeure'], time)
                        elif remark[:3] == 'S/D':
                            sheet.loc[x, 'grid_failure'] = add_time(sheet.loc[x, 'grid_failure'], time)
                        else:
                            sheet.loc[x, bd_factors[remark[:2]]] = add_time(sheet.loc[x, bd_factors[remark[:2]]], time)
                    else:
                        for f in factors:
                            if remark.find(f + ' -') != -1 or remark.find(f + '-') != -1:
                                sheet.loc[x, 'breakdown_hours'] = add_time(sheet.loc[x, 'breakdown_hours'], time)
                                if f == 'STS':
                                    sheet.loc[x, 'force_majeure'] = add_time(sheet.loc[x, 'force_majeure'], time)
                                elif f == 'S/D':
                                    sheet.loc[x, 'grid_failure'] = add_time(sheet.loc[x, 'grid_failure'], time)
                                else:
                                    sheet.loc[x, bd_factors[f]] = add_time(sheet.loc[x, bd_factors[f]], time)
                                break
                    remark = remark[index + 3:]

        sheet = timestamp_to_hours2(sheet, 'grid_failure')
        sheet = timestamp_to_hours2(sheet, 'force_majeure')
        sheet = timestamp_to_hours2(sheet, 'scheduled_services')
        sheet = timestamp_to_hours2(sheet, 'unscheduled_services')
        sheet = timestamp_to_hours2(sheet, 'breakdown_hours')

        sheet = sheet.reset_index().rename(columns={'Customer': 'customer_name',
                                                    'STATE ': 'state', 'SITE': 'site_name',
                                                    'WEC': 'wind_turbine_location_number',
                                                    'GENERATION': 'day_generation_kwh',
                                                    'MA ': 'machine_availability_percent',
                                                    'GIA': 'internal_grid_availability_percent',
                                                    'CF': 'plant_load_factor', 'REMARKS': 'breakdown_remark'})

        generation = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                            'state', 'site_name', 'wind_turbine_location_number',
                            'day_generation_kwh', 'operating_hours',
                            'machine_availability_percent', 'internal_grid_availability_percent',
                            'plant_load_factor', 'grid_failure', 'force_majeure',
                            'scheduled_services', 'unscheduled_services']]

        breakdown = sheet[['date', 'financial_year', 'customer_name', 'client_name',
                           'state', 'site_name', 'wind_turbine_location_number',
                           'breakdown_remark', 'breakdown_hours']]

        breakdown = breakdown[breakdown['breakdown_remark'] != '']

        gen_data = pd.concat([gen_data, generation], ignore_index=True)
        bd_data = pd.concat([bd_data, breakdown], ignore_index=True)

    return gen_data, bd_data


def consolidate():
    create_tables()

    gen_data = pd.DataFrame()
    bd_data = pd.DataFrame()

    gen_data, bd_data = get_inox_data(gen_data, bd_data)

    gen_data, bd_data = get_regen_data(gen_data, bd_data)

    gen_data, bd_data = get_senvion_data(gen_data, bd_data)

    gen_data, bd_data = get_suzlon_data(gen_data, bd_data)

    gen_data, bd_data = get_tswind_data(gen_data, bd_data)

    gen_data, bd_data = get_windworld_data(gen_data, bd_data)

    customers = {'D.J. Malpani': 'DJM', 'D J MALPANI': 'DJM', 'Giriraj Enterprises': 'GE', 'DJM': 'DJM',
                 'DJ Malpani Group': 'DJM',
                 'D J Malpani': 'DJM', 'NAKODA MACHINERY PVT. LTD.': 'NMPL', 'DJ Malpani': 'DJM',
                 'DJ Malpani - Palakkad': 'DJM',
                 'DJ Malpani - Sadla': 'DJM', 'DJ Malpani - Savarkundla': 'DJM', 'Giriraj Enterprises - Bagewadi': 'GE',
                 'IVY Ecoenergy India Private Ltd': 'IVY Ecoenergy India Private Ltd', 'Pravin Masalewale': 'PM',
                 'Hotel Golden Emerald': 'HGE'}

    if len(gen_data):
        gen_data['customer_name'] = [customers[x] for x in gen_data['customer_name']]
        gen_data.to_sql(con=db_connection, name='generation_mastersheet', if_exists='append', index=False)

    if len(bd_data):
        bd_data['customer_name'] = [customers[x] for x in bd_data['customer_name']]
        bd_data.to_sql(con=db_connection, name='breakdown_mastersheet', if_exists='append', index=False)

    cleanup('E:/SDA/Input DGR/')

    print('DGR Consolidated Successfully')
    print('-----------------------------', end='\n\n')
